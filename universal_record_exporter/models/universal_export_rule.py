# -*- coding: utf-8 -*-

import base64
import io
import re

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError
from odoo.tools import osutil


class UniversalExportRule(models.Model):
    _name = 'universal.export.rule'
    _description = 'Universal Export Rule'
    _rec_name = 'name'
    _order = 'name'

    name = fields.Char(string='Rule Name', required=True)

    model_id = fields.Many2one(
        'ir.model', string='Applies to', required=True, ondelete='cascade',
        help='The primary model this export rule targets.',
    )
    model_name = fields.Char(related='model_id.model', store=True, string='Model Technical Name')

    sub_model_id = fields.Many2one(
        'ir.model', string='Sub Model', ondelete='set null',
        help='The child model whose records form the detail rows (e.g. Order Lines).',
    )
    sub_model_name = fields.Char(related='sub_model_id.model', store=True)

    rel_field_id = fields.Many2one(
        'ir.model.fields', string='Relational Field', ondelete='set null',
        domain="[('model_id', '=', model_id), ('ttype', 'in', ['one2many', 'many2many'])]",
        help='The One2many / Many2many field on the primary model that links to the sub-model records.',
    )
    rel_field_name = fields.Char(related='rel_field_id.name', store=True)

    export_type = fields.Selection([
        ('excel', 'Export Excel'),
        ('text', 'CSV'),
        ('xml', 'XML'),
    ], string='Export Type', default='excel', required=True)

    active = fields.Boolean(default=True)
    export_file_name = fields.Char(string='Export File Name', default='export')

    field_line_ids = fields.One2many(
        'universal.export.line', 'export_id', string='Fields',
    )
    rel_field_line_ids = fields.One2many(
        'universal.export.rel.line', 'export_id', string='Relational Fields',
    )

    action_id = fields.Many2one(
        'ir.actions.server', string='Server Action', ondelete='set null', copy=False,
        readonly=True,
    )

    @api.onchange('rel_field_id')
    def _onchange_rel_field_id(self):
        IrModel = self.env['ir.model']
        for rule in self:
            if rule.rel_field_id and rule.rel_field_id.relation:
                rule.sub_model_id = IrModel.search([('model', '=', rule.rel_field_id.relation)], limit=1)
            else:
                rule.sub_model_id = False

    @api.onchange('model_id')
    def _onchange_model_id(self):
        for rule in self:
            rule.rel_field_id = False
            rule.sub_model_id = False

    @api.model
    def _get_sub_model_from_rel_field(self, rel_field):
        if not rel_field or not rel_field.relation:
            return self.env['ir.model']
        return self.env['ir.model'].search([('model', '=', rel_field.relation)], limit=1)

    @api.model
    def _set_sub_model_from_vals(self, vals):
        if vals.get('rel_field_id'):
            rel_field = self.env['ir.model.fields'].browse(vals['rel_field_id']).exists()
            vals['sub_model_id'] = self._get_sub_model_from_rel_field(rel_field).id or False
        elif 'rel_field_id' in vals:
            vals['sub_model_id'] = False

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            self._set_sub_model_from_vals(vals)
        return super().create(vals_list)

    def write(self, vals):
        vals = dict(vals)
        self._set_sub_model_from_vals(vals)
        return super().write(vals)

    @api.constrains('model_id', 'rel_field_id', 'sub_model_id')
    def _check_relation_configuration(self):
        for rule in self:
            if rule.rel_field_id:
                if rule.rel_field_id.model_id != rule.model_id:
                    raise ValidationError(_(
                        'The relational field must belong to the selected primary model.'
                    ))
                if rule.rel_field_id.ttype not in ('one2many', 'many2many'):
                    raise ValidationError(_('The relational field must be a One2many or Many2many field.'))
                expected_sub_model = self._get_sub_model_from_rel_field(rule.rel_field_id)
                if not expected_sub_model:
                    raise ValidationError(_('The sub model is required when a relational field is selected.'))
                if rule.sub_model_id and rule.sub_model_id != expected_sub_model:
                    raise ValidationError(_(
                        'The sub model "%(sub_model)s" must match the relation model "%(relation)s".',
                        sub_model=rule.sub_model_id.model,
                        relation=rule.rel_field_id.relation,
                    ))
            elif rule.sub_model_id:
                raise ValidationError(_('Select a relational field before setting a sub model.'))

    def action_add_in_model(self):
        self.ensure_one()
        if not self.env.user.has_group('base.group_system'):
            raise UserError(_('Only Settings administrators can add export actions to models.'))
        if not self.model_id:
            raise UserError(_('Please select a model first.'))
        self._validate_configuration()

        action_name = 'Export %s' % self.name
        code = (
            "config = env['universal.export.rule'].browse(%d)\n"
            "action = config.do_export(env.context.get('active_ids') or [])"
            % self.id
        )
        export_group = self.env.ref('base.group_allow_export')

        if self.action_id:
            self.action_id.sudo().write({
                'name': action_name,
                'model_id': self.model_id.id,
                'binding_model_id': self.model_id.id,
                'binding_view_types': 'list,form',
                'groups_id': [(6, 0, export_group.ids)],
                'code': code,
            })
        else:
            server_action = self.env['ir.actions.server'].sudo().create({
                'name': action_name,
                'model_id': self.model_id.id,
                'binding_model_id': self.model_id.id,
                'binding_type': 'action',
                'binding_view_types': 'list,form',
                'groups_id': [(6, 0, export_group.ids)],
                'state': 'code',
                'code': code,
            })
            self.action_id = server_action

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Export Action Added'),
                'message': _(
                    'The export action "%s" is now available in the Action menu of %s.'
                ) % (action_name, self.model_id.name),
                'type': 'success',
                'sticky': False,
            },
        }

    def action_remove_from_model(self):
        """Remove the bound server action from the target model."""
        self.ensure_one()
        if not self.env.user.has_group('base.group_system'):
            raise UserError(_('Only Settings administrators can remove export actions from models.'))
        if self.action_id:
            self.action_id.sudo().unlink()
            self.action_id = False
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Export Action Removed'),
                'message': _('The export action has been removed from the model.'),
                'type': 'warning',
                'sticky': False,
            },
        }

    def unlink(self):
        actions = self.mapped('action_id').sudo()
        result = super().unlink()
        actions.unlink()
        return result

    def do_export(self, active_ids):
        self.ensure_one()
        if not (self.env.is_admin() or self.env.user.has_group('base.group_allow_export')):
            raise UserError(_('You do not have the rights to export data. Please contact an administrator.'))
        if not active_ids:
            raise UserError(_('No records selected for export.'))
        self._validate_configuration()

        records = self.env[self.model_id.model].browse(active_ids).exists()
        if not records:
            raise UserError(_('No accessible records were found for export.'))
        records.check_access('read')

        if self.export_type == 'excel':
            return self._export_excel(records)
        elif self.export_type == 'text':
            return self._export_text(records)
        elif self.export_type == 'xml':
            return self._export_xml(records)
        else:
            raise UserError(_('Unsupported export type: %s') % self.export_type)

    def _validate_configuration(self):
        self.ensure_one()
        if not self.field_line_ids and not self.rel_field_line_ids:
            raise UserError(_('Add at least one field before exporting.'))
        if self.rel_field_id and not self.rel_field_line_ids:
            raise UserError(_('Add at least one related model field or remove the relational field.'))
        if self.rel_field_line_ids and not self.rel_field_id:
            raise UserError(_('Select a relational field before adding related model fields.'))
        self._check_relation_configuration()

    def _get_file_stem(self):
        stem = (self.export_file_name or self.name or 'export').strip() or 'export'
        for extension in ('.xlsx', '.csv', '.txt', '.xml'):
            if stem.lower().endswith(extension):
                stem = stem[:-len(extension)]
                break
        return osutil.clean_filename(stem.strip() or 'export') or 'export'

    def _get_file_name(self, extension):
        return '%s.%s' % (self._get_file_stem(), extension)

    def _get_sheet_title(self):
        title = re.sub(r'[\[\]\:\*\?\/\\]', ' ', self.name or 'Export').strip()
        return (title or 'Export')[:31]

    def _selection_label(self, record, field_id, value):
        try:
            field_obj = record._fields.get(field_id.name)
            if not field_obj:
                return value
            selection = field_obj.selection
            if callable(selection):
                selection = selection(record)
            return dict(selection).get(value, value)
        except Exception:
            return value

    def _format_value(self, value, field_id=False, record=False):
        if value is False or value is None:
            return ''
        if hasattr(value, '_name'):
            return ', '.join(value.mapped('display_name')) if value else ''
        if field_id and record and field_id.ttype == 'selection':
            value = self._selection_label(record, field_id, value)
        return str(value)

    def _get_value(self, record, field_id, rel_field_id=False):
        """
        Read a field value from *record*.

        * If *rel_field_id* is set the value is first read as a relational
          object and then the sub-field is read from that object.
        * Selection fields are translated to their human-readable label.
        """
        if not field_id:
            return ''

        field_name = field_id.name
        if field_name not in record._fields:
            return ''

        value = getattr(record, field_name, False)

        if value is False or value is None:
            return ''

        if rel_field_id:
            rel_name = rel_field_id.name
            if not hasattr(value, '_name'):
                return ''
            values = [
                self._format_value(getattr(item, rel_name, False), rel_field_id, item)
                for item in value
            ]
            return ', '.join(item for item in values if item)

        return self._format_value(value, field_id, record)

    def _build_rows(self, records):
        """
        Return (headers, data_rows).

        Each data_row is a flat list aligned with headers.
        If a sub-model is configured every header-record combination is
        exploded into one row per sub-record.
        """
        main_labels = [
            line.label or line.field_id.field_description
            for line in self.field_line_ids
        ]
        rel_labels = [
            line.label or line.field_id.field_description
            for line in self.rel_field_line_ids
        ]
        headers = main_labels + rel_labels

        rows = []
        use_sub = bool(self.rel_field_id and self.rel_field_line_ids)

        for record in records:
            main_vals = [
                self._get_value(record, line.field_id, line.rel_field_id)
                for line in self.field_line_ids
            ]

            if use_sub:
                sub_records = getattr(record, self.rel_field_id.name, [])
                if sub_records:
                    for sub in sub_records:
                        rel_vals = [
                            self._get_value(sub, line.field_id, line.rel_field_id)
                            for line in self.rel_field_line_ids
                        ]
                        rows.append(main_vals + rel_vals)
                else:
                    rows.append(main_vals + [''] * len(rel_labels))
            else:
                rows.append(main_vals)

        return headers, rows

    def _export_excel(self, records):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError(_(
                'The Python library "openpyxl" is required for Excel export.\n'
                'Install it with: pip install openpyxl'
            ))

        headers, rows = self._build_rows(records)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._get_sheet_title()

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(border_style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        ws.row_dimensions[1].height = 20

        alt_fill = PatternFill(start_color='EBF0FB', end_color='EBF0FB', fill_type='solid')
        data_align = Alignment(vertical='center', wrap_text=False)

        for row_idx, row in enumerate(rows, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_align
                cell.border = border
                if fill:
                    cell.fill = fill

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read()).decode()
        file_name = self._get_file_name('xlsx')

        return self._create_attachment_action(
            file_name, file_data,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _export_text(self, records):
        import csv

        headers, rows = self._build_rows(records)

        output = io.StringIO()
        writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        writer.writerows(rows)

        file_data = base64.b64encode(output.getvalue().encode('utf-8')).decode()
        file_name = self._get_file_name('csv')

        return self._create_attachment_action(file_name, file_data, 'text/csv;charset=utf-8')

    def _export_xml(self, records):
        import xml.etree.ElementTree as ET
        from xml.dom import minidom

        headers, rows = self._build_rows(records)

        root = ET.Element('records')
        root.set('model', self.model_id.model)
        root.set('export', self.name)

        for row in rows:
            rec_el = ET.SubElement(root, 'record')
            for header, value in zip(headers, row):
                field_el = ET.SubElement(rec_el, 'field')
                field_el.set('name', header)
                field_el.text = str(value) if value else ''

        xml_str = minidom.parseString(
            ET.tostring(root, encoding='unicode')
        ).toprettyxml(indent='  ')

        file_data = base64.b64encode(xml_str.encode('utf-8')).decode()
        file_name = self._get_file_name('xml')

        return self._create_attachment_action(file_name, file_data, 'application/xml')

    def _create_attachment_action(self, file_name, file_data, mimetype):
        attachment = self.env['ir.attachment'].sudo().create({
            'name': file_name,
            'type': 'binary',
            'datas': file_data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': mimetype,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'new',
        }
